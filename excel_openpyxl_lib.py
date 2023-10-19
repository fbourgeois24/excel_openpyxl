# Documentation
# - https://openpyxl.readthedocs.io/en/stable/tutorial.html

from openpyxl import load_workbook, Workbook
from openpyxl import styles as excel_style
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from openpyxl.styles.borders import Border, Side
# import logging as log

class excel_file():

	def __init__(self, filepath, next_line=1, data_only=False):
		"""
			filepath est le chemin complet du fichier
			next_line est la ligne suivante à laquelle écrire, par défaut, la première
		"""
		self.filepath = filepath
		self.next_line = next_line
		self.data_only = data_only
		self.save = False # Sauver à la fermeture


	def create_workbook(self, sheet_name=None):
		self.wb = Workbook()
		# S'il faut donner un nom à la feuille
		if sheet_name is not None:
			self.ws = self.wb.active
			self.ws.title = sheet_name
		self.wb.save(self.filepath)

	def create_sheet(self, sheet_name):
		""" Créer une feuille """
		self.ws[sheet_name] = self.wb.create_sheet(sheet_name)

	def open(self, data_only=False):
		""" Ouverture du document """
		# On défini le data_only
		self.data_only=max(data_only, self.data_only)
		self.wb = load_workbook(self.filepath, data_only=self.data_only)
		self.ws = {}
		for sheet in self.wb.sheetnames:
			self.ws[sheet] = self.wb[sheet]

	def save(self):
		""" Sauver le document """
		self.wb.save(self.filepath)

	def close(self, save=False):
		if max(save, self.save):
			self.save()
		self.wb.close()

	def __enter__(self):
		# Ouverture avec WITH
		self.open()

	def __exit__(self, *args, **kwargs):
		""" Fermeture après WITH 
			Pas de sauvergarde de la feuille si utilisation avec data_only
		"""
		self.close(save=self.save)

	def read(self, range, sheet=0):
		""" Lire une valeur dans le classeur """
		cell_range = self.ws[self.find_sheet(sheet)][range]
		if not ":" in range:
			# Si une seule cellule on renvoie la valeur
			return cell_range.value
		else:
			# Si plusiers cellules
			cell_values = []
			for line in cell_range:
				line_values = []
				for cell in line:
					print(cell.value)
					line_values.append(cell.value)
				cell_values.append(line_values)
			return cell_values

	def write(self, data, sheet=0, from_cell=1, from_line=None):
		""" Ecriture dans le fichier, 
			Données à écrire:
				- Si données n'est pas une liste ou un tuple, on écrit une seule case
				- Si données est une liste ou un tuple à un seul niveau, on écrit une ligne
				- Si données est une liste ou un tuple à deux niveaux, on écrit plusieurs lignes
			Ligne à écrire (une ligne libre est une ligne dont la première cellule est vide):
				- si from_line est null on écrit à la ligne suivante enregistrée dans self.next_line
				- si line est à first_free, on cherche la première ligne libre
				- si line est à next_free, on cherche la ligne libre suivante après self.next_line
		"""
		# On active la sauvegarde à la fermeture
		self.save = True

		# Si on accède à la feuille par son indice, on récupère son nom
		sheet = self.find_sheet(sheet)

		# On défini la ligne à écrire
		if from_line is None:	
			from_line = self.next_line
		elif from_line == "first_free":
			while self.ws[sheet].cell(1, 1).value is not None:
				from_line += 1
		elif from_line == 'next_free':
			while self.ws[sheet].cell(self.next_line, 1).value is not None:
				from_line += 1
		elif str(from_line).isdigit():
			pass
		else:
			raise TypeError("Valeur de ligne à écrire incorrecte")

		# On défini la taille des données à écrire
		# data_size est une liste qui contiendra les deux dimentions des données à écrite [ligne, colonne]
		if type(data) not in (tuple, list):
			# Cellule unique
			data_size = [1,1]
			self.ws[sheet].cell(from_line, from_cell).value = data
		elif type(data[0]) not in (tuple, list):
			# On défini si c'est une liste à deux dimentions en regardant le premier élément de la liste
			# Liste à une dimention
			data_size = [1, len(data)]
			for col in range(1, data_size[1]+1):
				self.ws[sheet].cell(from_line, col + from_cell).value = data[col-1]
		else:
			# Liste à deux dimentions
			data_size = [len(data[0]), len(data)]
			# On écrit les données en fonction des tailles définies 
			for line in range(1, data_size[1]+1):
				for col in range(1, data_size[0]+1):
					self.ws[sheet].cell(line + from_line, col + from_cell).value = data[line-1][col-1]


	def append(self, data, sheet=0):
		""" Ajout d'une ou plusieurs lignes directement à la suite
			Data doit obligatoirement être un tuple ou une liste même s'il n'y a qu'un seul élément
		"""
		# Si on accède à la feuille par son indice, on récupère son nom
		sheet = self.find_sheet(sheet)

		if type(data) not in (list, tuple):
			raise TypeError("Les données doivent être une liste ou un tuple à une ou deux dimentions")
		if type(data[0]) in (list, tuple):
			# Si data a deux dimentions, on boucle car la mathode append ne les gère pas
			for item in data:
				self.ws[sheet].append(item)
		else:
			self.ws[sheet].append(data)



	def auto_fit(self, sheet=0, column_start=1, column_end=100, row_start=1, row_end=100):
		""" Redimentionnement des colonnes à la taille du contenu
			Par défaut pour les 100 premières colonnes et 100 premières lignes
		"""
		# Si on accède à la feuille par son indice, on récupère son nom
		sheet = self.find_sheet(sheet)

		for column in range(column_start, column_end + 1):
			max_width = 10
			for row in range(row_start, row_end + 1):
				width = len(str(self.ws[sheet].cell(row, column).value)) * 1.1
				if width > max_width:
					max_width = width
			self.ws[sheet].column_dimensions[get_column_letter(column)].width = max_width


	def color_row(self, row_number, color, fill_type = "solid", sheet=0):
		""" Colorer une ligne """
		
		# Si on accède à la feuille par son indice, on récupère son nom
		sheet = self.find_sheet(sheet)

		for rows in self.ws[sheet].iter_cols(min_row=row_number, max_row=row_number, min_col=None, max_col=None):
			for cell in rows:
				cell.fill = PatternFill(start_color=color, end_color=color, fill_type = fill_type)

	def color_column(self, col_number, color, fill_type = "solid", sheet=0, min_row=1, max_row=1000):
		""" Colorer une colonne """
		
		# Si on accède à la feuille par son indice, on récupère son nom
		sheet = self.find_sheet(sheet)

		for col in self.ws[sheet].iter_cols(min_row=1, max_row=max_row, min_col=col_number, max_col=col_number):
			for cell in col:
				cell.fill = PatternFill(start_color=color, end_color=color, fill_type = fill_type)


	def find_sheet(self, sheet):
		""" Trouver une feuille si recherche par le nom ou l'indice 
			Renvoie le nom de la feuille
		"""

		# Si on accède à la feuille par son indice, on récupère son nom
		if type(sheet) == int:
			sheet_name = tuple(self.ws)[sheet]
		elif type(sheet) == str:
			sheet_name = sheet
		else:
			raise ValueError("sheet doit être une string ou un entier")

		return sheet_name


	def border(self, range, sheet=0):
		""" Ajouter des bordures aux cellules
		"""

		sheet = self.find_sheet(sheet)
		cell_range = self.ws[self.find_sheet(sheet)][range]
		thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

		for line in cell_range:
			for cell in line:
				cell.border = thin_border







